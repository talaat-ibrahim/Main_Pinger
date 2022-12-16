from icmplib import ping
import ascii_train
import os
import pathlib
import openpyxl
print(ascii_train.train("This Code Provided By En.Talaat Ibrahim"))
file = openpyxl.load_workbook("src.xlsx")
sheet = file["Sheet1"]
rows = sheet.max_row
columns = sheet.max_column
for row  in range(1,rows):
    for column in range(1, columns):
        getDataCell = (sheet.cell(row=row, column=column).value)
        row +1
        result = ping(address=str(getDataCell), count=2,privileged=False, interval=0.2, timeout=1)
        print(result)
        if (result.is_alive ):
            outOnline = (f"[ {getDataCell} ] -UP|| Packet_loss= {result.packet_loss*100}%")
            sheet.cell(row=row, column=2, value=outOnline)
        else:
            outoffline =  (f"[ {getDataCell} ] -DOWN|| Packet_loss= {result.packet_loss*100}%")
            sheet.cell(row=row, column=2, value=outoffline)
    file.save("src.xlsx")
print("The Tsak is Done || PLZ Remmeber me with Good || EN.Talaat ".upper)
